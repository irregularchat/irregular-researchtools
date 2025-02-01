# /pages/frameworks/ACH.py
import streamlit as st
from dotenv import load_dotenv
from utilities.utils_openai import chat_gpt  # If you want AI to assist
import pandas as pd
from io import BytesIO
import openpyxl

load_dotenv()

def ach_page():
    st.title("Analysis of Competing Hypotheses (ACH)")

    st.write("""
    **Analysis of Competing Hypotheses** (ACH) helps evaluate multiple competing explanations
    or hypotheses for the observed data. This approach was popularized by Richards Heuer at the CIA.
    
    **Steps**:
    1. List possible Hypotheses.
    2. List key Pieces of Evidence or Arguments.
    3. For each piece of evidence, evaluate how Consistent, Inconsistent, or Neutral it is with each hypothesis.
    4. Identify which hypothesis has the fewest inconsistencies.
    5. Refine or add evidence as needed to challenge your initial assumptions.
    """)

    st.markdown("---")

    #
    # Step 1: Hypotheses
    #
    st.subheader("Step 1: List Hypotheses")

    if "ach_hypotheses" not in st.session_state:
        st.session_state["ach_hypotheses"] = ""

    def ai_suggest_hypotheses():
        """
        Call AI to propose 3–5 example hypotheses from a senior intelligence analyst perspective.
        Returns semicolon-separated text for each hypothesis.
        """
        try:
            existing_hypotheses = st.session_state["ach_hypotheses"]

            system_msg = {
                "role": "system",
                "content": (
                    "You are an AI with expertise in intelligence analysis, using the Analysis of Competing Hypotheses (ACH) approach. "
                    "You will propose a brief set of well-reasoned hypotheses, returning them as semicolon-limited text. "
                    "Be concise, yet ensure each hypothesis distinctly addresses possible underlying causes, scenarios, or factors."
                )
            }

            user_msg = {
                "role": "user",
                "content": (
                    f"Current hypotheses: {existing_hypotheses}\n\n"
                    "As a senior analyst applying ACH, propose 3–5 new or refined hypotheses (semicolon separated) that could explain or address "
                    "this scenario. Consider different angles, assumptions, and any gaps in the existing hypotheses. "
                    "Focus on plausible scenarios, mention potential uncertainties, and ensure each proposed hypothesis is distinct."
                )
            }

            resp = chat_gpt([system_msg, user_msg], model="gpt-4o-mini")
            
            # Split the response by semicolons and format into new lines
            formatted_resp = "\n".join(resp.split(';'))
            return formatted_resp
        except Exception as e:
            st.error(f"AI error: {e}")
            return ""

    col_hyp_left, col_hyp_right = st.columns([3,1])
    with col_hyp_left:
        st.session_state["ach_hypotheses"] = st.text_area(
            "Hypotheses (semicolon or newline separated)",
            value=st.session_state["ach_hypotheses"],
            placeholder="Hypothesis A\nHypothesis B\nHypothesis C",
            height=100
        )
    with col_hyp_right:
        if st.button("AI: Generate Options for Hypotheses"):
            text = ai_suggest_hypotheses()
            if text:
                st.session_state["ach_hypotheses"] += "\n" + text
                st.experimental_rerun()

    st.markdown("---")

    #
    # Step 2: Evidence / Arguments
    #
    st.subheader("Step 2: List Key Pieces of Evidence or Arguments (Assign Weights)")

    if "ach_evidence" not in st.session_state:
        st.session_state["ach_evidence"] = ""

    # Initialize a dict to store weights for each piece of evidence.
    if "ach_evidence_weights" not in st.session_state:
        st.session_state["ach_evidence_weights"] = {}

    def ai_suggest_evidence():
        """Call AI to propose 3–5 example pieces of evidence or arguments for the scenario."""
        try:
            existing_hypotheses = st.session_state["ach_hypotheses"]
            system_msg = {"role": "system", "content": "You are an AI that suggests evidence for ACH."}
            user_msg = {
                "role": "user",
                "content": (
                    f"Based on the current hypotheses: {existing_hypotheses}, "
                    "suggest 3 to 5 pieces of evidence or arguments (semicolon separated) that are relevant to these hypotheses."
                )
            }
            resp = chat_gpt([system_msg, user_msg], model="gpt-4o-mini")
            
            # Split the response by semicolons and format into new lines
            formatted_resp = "\n".join(resp.split(';'))
            return formatted_resp
        except Exception as e:
            st.error(f"AI error: {e}")
            return ""

    col_evid_left, col_evid_right = st.columns([3,1])
    with col_evid_left:
        st.session_state["ach_evidence"] = st.text_area(
            "Evidence / Arguments",
            value=st.session_state["ach_evidence"],
            placeholder="Evidence #1\nEvidence #2\netc.",
            height=100
        )
    with col_evid_right:
        if st.button("AI: Generate Options for Evidence to Consider"):
            text = ai_suggest_evidence()
            if text:
                st.session_state["ach_evidence"] += "\n" + text
                st.experimental_rerun()

    st.markdown("---")

    #
    # Step 2b: Assign Weights
    #
    st.write("Assign a weight for each piece of evidence (e.g., 1 = least significant / Opinion, 12 = extremely significant/ Fact).")
    def parse_list(input_text):
        # Split by semicolon, then newline
        parts = input_text.split(';')
        result = []
        for part in parts:
            result.extend([x.strip() for x in part.split('\n') if x.strip()])
        return result

    evidence_list = parse_list(st.session_state["ach_evidence"])

    # Define the logarithmic scale
    weight_options = [1, 3, 5, 8, 12]

    def format_weight_option(weight):
        if weight == 1:
            return "1 (Least significant, opinion-based)"
        elif weight == 12:
            return "12 (Extremely significant, factual)"
        else:
            return str(weight)

    for ev in evidence_list:
        if ev not in st.session_state["ach_evidence_weights"]:
            # Default weight set to 1
            st.session_state["ach_evidence_weights"][ev] = 1

    for ev in evidence_list:
        st.session_state["ach_evidence_weights"][ev] = st.selectbox(
            label=f"Select weight for: {ev}",
            options=weight_options,
            index=weight_options.index(st.session_state["ach_evidence_weights"][ev]),
            format_func=format_weight_option,
            key=f"{ev}_weight_selectbox"
        )

    st.markdown("---")

    #
    # Step 3: Evaluate Consistency of Each Piece of Evidence with Each Hypothesis
    #
    st.subheader("Step 3: Evaluate Consistency")

    hypotheses_list = parse_list(st.session_state["ach_hypotheses"])

    # We'll keep a data structure for the consistency matrix
    if "ach_matrix" not in st.session_state:
        st.session_state["ach_matrix"] = {}

    # Ensure session_state has a record for each (evid, hyp) pair
    for e in evidence_list:
        for h in hypotheses_list:
            key_ = (e, h)
            if key_ not in st.session_state["ach_matrix"]:
                st.session_state["ach_matrix"][key_] = "Neutral"

    if hypotheses_list and evidence_list:
        for ev in evidence_list:
            st.markdown(f"**Evidence**: {ev}")
            columns = st.columns(len(hypotheses_list))
            for i, hyp in enumerate(hypotheses_list):
                old_val = st.session_state["ach_matrix"][(ev, hyp)]
                new_val = columns[i].selectbox(
                    label=f"{hyp}",
                    options=["Consistent", "Inconsistent", "Neutral"],
                    index=["Consistent", "Inconsistent", "Neutral"].index(old_val),
                    key=f"{ev}_{hyp}_selectbox"
                )
                st.session_state["ach_matrix"][(ev, hyp)] = new_val
            st.write("---")
    else:
        st.info("Add at least one hypothesis and one piece of evidence to compare them.")

    st.markdown("---")

    #
    # Step 4: Weighted Inconsistency Counts
    #
    st.subheader("Step 4: Assess Weighted Inconsistencies / Conflicts")
    st.write("""
    In this weighted approach, each piece of evidence you marked "Inconsistent" applies its assigned weight
    negatively, "Consistent" positively, and "Neutral" as zero to the hypothesis' total score.
    """)

    if hypotheses_list and evidence_list:
        # Initialize scores
        weighted_score = {h: 0.0 for h in hypotheses_list}
        consistency_counts = {h: {"Consistent": 0, "Inconsistent": 0, "Neutral": 0} for h in hypotheses_list}

        for e in evidence_list:
            w = st.session_state["ach_evidence_weights"][e]
            for h in hypotheses_list:
                val = st.session_state["ach_matrix"][(e, h)]
                consistency_counts[h][val] += 1
                if val == "Consistent":
                    weighted_score[h] += w
                elif val == "Inconsistent":
                    weighted_score[h] -= w
                # Neutral contributes 0, so no change needed

        # Prepare display
        sorted_hyps = sorted(weighted_score.items(), key=lambda x: x[1], reverse=True)
        
        # Display the tentative conclusion first
        best_hyp = sorted_hyps[0][0] if sorted_hyps else None
        st.success(f"Tentative Conclusion: Hypothesis with the highest weighted score is: **{best_hyp}**")
        
        st.write("**Weighted Scores and Counts** (higher is better):")
        
        for hyp, score in sorted_hyps:
            counts = consistency_counts[hyp]
            
            st.markdown(f"### {hyp}")
            st.markdown(f"- **Score**: {score:.2f}")
            st.markdown(f"- **Consistent**: {counts['Consistent']}")
            st.markdown(f"- **Inconsistent**: {counts['Inconsistent']}")
            st.markdown(f"- **Neutral**: {counts['Neutral']}")
            st.markdown("---")

        # Create two columns for the buttons
        col1, col2 = st.columns(2)

        # Devil's Advocate button
        with col1:
            if st.button("Devil's Advocate: Challenge Hypotheses"):
                devils_advocate_response = ai_devils_advocate(
                    hypotheses_list, 
                    evidence_list, 
                    weighted_score, 
                    consistency_counts
                )
                st.write(devils_advocate_response)

        # Export button
        with col2:
            # Correct session state retrieval
            hypotheses = st.session_state.get('ach_hypotheses', [])
            evidence = st.session_state.get('ach_evidence', [])
            matrix = st.session_state.get('ach_matrix', {})
            weights = st.session_state.get('ach_evidence_weights', {})

            # Always show the export button
            if st.button('Export to Excel'):
                try:
                    # Export
                    excel_data = export_to_excel(hypotheses, evidence, matrix, weights)
                    
                    if excel_data:
                        st.download_button(
                            label="📥 Download Excel file",
                            data=excel_data,
                            file_name='ACH_matrix.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                except Exception as e:
                    st.error(f"Export failed: {str(e)}")

    else:
        st.info("No matrix to analyze. Add Hypotheses and Evidence first.")

    st.markdown("---")
    st.info("""
**Done!**  
Use these steps to refine or add new evidence. Weighting helps you see which evidence
has the most impact on your analysis, offering a more nuanced view of each hypothesis.
""")

def ai_devils_advocate(hypotheses, evidence, weighted_score, consistency_counts):
    """Call AI to provide counterarguments and perspectives."""
    try:
        # Prepare a detailed string of inconsistencies
        inconsistencies_detail = "\n".join(
            f"- {h}: {', '.join([e for e, val in consistency_counts[h].items() if val == 'Inconsistent'])}"
            for h in hypotheses
        )

        system_msg = {
            "role": "system",
            "content": (
                "You are an AI tasked with providing a devil's advocate perspective on hypotheses. "
                "Offer professional counterarguments, questions to consider, and alternative perspectives. "
                "Don't go out of your way to be a contrarian; the feedback should be helpful."
            )
        }
        user_msg = {
            "role": "user",
            "content": (
                f"Hypotheses: {hypotheses}\n"
                f"Evidence: {evidence}\n"
                f"Weighted Scores: {weighted_score}\n"
                f"Inconsistencies:\n{inconsistencies_detail}\n\n"
                "Provide counterarguments, questions, and perspectives to challenge the current hypothesis selection."
            )
        }
        resp = chat_gpt([system_msg, user_msg], model="gpt-4o-mini")
        return resp
    except Exception as e:
        st.error(f"AI error: {e}")
        return ""

def export_to_excel(hypotheses_list, evidence_list, ach_matrix, evidence_weights):
    """Export the ACH matrix to an Excel file."""
    try:
        # Ensure hypotheses_list is a list
        if isinstance(hypotheses_list, str):
            hypotheses_list = hypotheses_list.split('\n')  # Split by newline if it's a single string

        # Ensure evidence_list is a list
        if isinstance(evidence_list, str):
            evidence_list = evidence_list.split('\n')  # Split by newline if it's a single string

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'ACH Matrix'
        
        # Style for headers
        header_style = openpyxl.styles.NamedStyle(name='header')
        header_style.font = openpyxl.styles.Font(bold=True)
        header_style.fill = openpyxl.styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Write headers
        headers = ['Evidence', 'Weight'] + hypotheses_list
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = header_style
            
        # Write data
        for row_idx, evidence in enumerate(evidence_list, 2):  # Start from row 2
            # Evidence and weight
            ws.cell(row=row_idx, column=1, value=str(evidence))
            ws.cell(row=row_idx, column=2, value=float(evidence_weights.get(evidence, 0)))
            
            # Consistency values
            for col_idx, hypothesis in enumerate(hypotheses_list, 3):  # Start from column 3
                key = (evidence, hypothesis)
                value = ach_matrix.get(key, "N/A")  # Use "N/A" for missing data
                ws.cell(row=row_idx, column=col_idx, value=str(value))
        
        # Calculate total scores
        total_scores = {h: 0 for h in hypotheses_list}
        for (evidence, hypothesis), consistency in ach_matrix.items():
            if hypothesis in total_scores:  # Ensure hypothesis is in total_scores
                weight = evidence_weights.get(evidence, 0)
                if consistency == 'Consistent':
                    total_scores[hypothesis] += weight
                elif consistency == 'Inconsistent':
                    total_scores[hypothesis] -= weight
        
        # Write total scores
        total_row = len(evidence_list) + 2
        ws.cell(row=total_row, column=1, value='Total Score')
        for col_idx, hypothesis in enumerate(hypotheses_list, 3):
            ws.cell(row=total_row, column=col_idx, value=total_scores.get(hypothesis, 0))
        
        # Auto-adjust columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
            
        # Add borders
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=total_row, 
                              min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

    except Exception as e:
        import traceback
        error_msg = f"Export error: {str(e)}\n{traceback.format_exc()}"
        st.error(error_msg)
        return None

def main():
    ach_page()

if __name__ == "__main__":
    main()