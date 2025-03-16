import io
from io import BytesIO

import pytest
import pandas as pd
from docx import Document
from openpyxl import load_workbook, Workbook

from utilities.helpers import (
    remove_markdown,
    create_docx_document,
    export_ach_matrix_to_excel,
    apply_excel_styling,
)

def test_remove_markdown():
    # Test removal of bold formatting
    input_text = "**Bold Text**"
    expected_output = "Bold Text"
    assert remove_markdown(input_text) == expected_output

    # Test removal of italic formatting (using asterisks)
    input_text = "*Italic Text*"
    expected_output = "Italic Text"
    assert remove_markdown(input_text) == expected_output

    # Test removal of combined markdown elements: header, link, and bold.
    input_text = "# Header Title\nThis is a [link](https://example.com) and **bold** text"
    expected_output = "Header Title\nThis is a link and bold text"
    assert remove_markdown(input_text) == expected_output

    # Test empty input
    assert remove_markdown("") == ""

def test_create_docx_document():
    title = "Test Document"
    sections = {
        "Section 1": "**Bold Section Content**",
        "Section 2": "Content with *italic* text",
    }
    docx_io = create_docx_document(title, sections)
    
    # Ensure the returned object is a BytesIO instance and is non-empty.
    assert isinstance(docx_io, BytesIO)
    contents = docx_io.getvalue()
    assert len(contents) > 0

    # Use the Document API to read back the DOCX content.
    document = Document(BytesIO(contents))
    paragraphs = [p.text for p in document.paragraphs]

    # Check that the title appears as a heading.
    assert title in paragraphs

    # Check that section headings appear in the document.
    for section in sections.keys():
        assert section in paragraphs

def test_export_ach_matrix_to_excel():
    # Prepare dummy ACH data.
    hypotheses_list = ["Hypothesis A", "Hypothesis B"]
    evidence_list = ["Evidence 1", "Evidence 2"]
    ach_matrix = {
        ("Evidence 1", "Hypothesis A"): "Consistent",
        ("Evidence 1", "Hypothesis B"): "Inconsistent",
        ("Evidence 2", "Hypothesis A"): "Neutral",
        ("Evidence 2", "Hypothesis B"): "Consistent",
    }
    evidence_weights = {"Evidence 1": 5, "Evidence 2": 3}

    excel_bytes = export_ach_matrix_to_excel(
        hypotheses_list, evidence_list, ach_matrix, evidence_weights
    )
    
    # Check that the result is a bytes object and is non-empty.
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0

    # Load the Excel file using openpyxl to inspect its contents
    excel_file = BytesIO(excel_bytes)
    wb = load_workbook(excel_file)
    sheet_name = "ACH Matrix"
    assert sheet_name in wb.sheetnames
    ws = wb[sheet_name]
    
    # Verify headers: should include "Evidence", "Weight" and each hypothesis.
    headers = [cell.value for cell in ws[1]]
    expected_headers = ["Evidence", "Weight"] + hypotheses_list
    for header in expected_headers:
        assert header in headers
        
    # Check row count: one header row plus one row per piece of evidence.
    assert ws.max_row == len(evidence_list) + 1

def test_apply_excel_styling():
    # Create a dummy workbook and worksheet.
    wb = Workbook()
    ws = wb.active
    headers = ["A", "B", "C"]
    # Append header row and one data row.
    ws.append(headers)
    ws.append(["Value1", "Value2", "Value3"])
    
    # Apply styling to the worksheet.
    apply_excel_styling(ws, headers, total_row=ws.max_row)
    
    # Check that header cells now have a border style.
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        # The border object is expected to have non-default side settings.
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"
        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style == "thin"
    
    # Verify that column widths were adjusted.
    for column_cells in ws.columns:
        width = ws.column_dimensions[column_cells[0].column_letter].width
        header_text = str(column_cells[0].value)
        # Expected minimal width is length of header text plus padding.
        assert width >= len(header_text) + 2

@pytest.mark.parametrize("input_text,expected", [
    ("", ""),  # Empty string
    (None, ""),  # None input
    ("Plain text with no markdown", "Plain text with no markdown"),  # No markdown
    ("```code block```", "code block"),  # Code blocks
    ("> Blockquote", "Blockquote"),  # Blockquotes
    ("1. First\n2. Second", "1. First\n2. Second"),  # Lists
    ("---", ""),  # Horizontal rules
])
def test_remove_markdown_edge_cases(input_text, expected):
    assert remove_markdown(input_text) == expected

if __name__ == "__main__":
    pytest.main() 